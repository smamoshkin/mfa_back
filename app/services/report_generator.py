# app/services/report_generator.py

from typing import Dict, Any, List
from datetime import datetime
from sqlalchemy.orm import Session
from sqlalchemy import text
import pandas as pd
from io import BytesIO
import logging

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


class DynamicReport:
    """Генератор динамических отчетов с формулами в Excel"""
    
    # Соответствие колонок БД и русских названий для отчета
    COLUMN_MAPPING = {
        # Базовые данные
        'product_name': 'Название товара',
        'sku': 'Артикул',
        
        # Продажи
        'quantity_sold': 'Продано, шт.',
        'revenue': 'Выручка WB, ₽',
        'seller_payout': 'Перечислено продавцу, ₽',
        
        # Расходы
        'tax': 'Налог, ₽',
        'payout_after_tax': 'Перечисление после налога, ₽',
        'storage_fee': 'Хранение, ₽',
        'regular_deduction': 'Удержание, ₽',
        'dzhem_deduction': 'Джем, ₽',
        'delivery_rub': 'Логистика, ₽',
        'penalty': 'Штрафы, ₽',
        'acceptance': 'Приемка, ₽',
        
        # Маржа и себестоимость
        'margin': 'Маржа, ₽',
        'cost_per_unit': 'Себестоимость единицы, ₽',
        'total_cost': 'Общая себестоимость, ₽',
        
        # Возвраты
        'return_quantity': 'Возвраты, шт.',
        'return_revenue': 'Возвраты, ₽',
    }
    
    def __init__(self, db_session: Session):
        self.db = db_session
    
    def generate_excel_report(self, filters: Dict[str, Any], tenant_id: int) -> BytesIO:
        """
        Основной метод: генерирует Excel отчет с формулами
        Только ОДИН запрос к БД
        Структура:
        1. Заголовок (строки 1-3)
        2. Вертикальная часть с формулами (строки 5-30, колонки A-B)
        3. Горизонтальная таблица с товарами (под вертикальной частью)
        """
        try:
            # 1. ОДИН запрос к БД для получения всех данных
            raw_df = self._get_report_data(filters, tenant_id)
            
            if raw_df.empty:
                logger.error("Нет данных для отчета")
                raise ValueError("Нет данных для формирования отчета")
            
            # 2. Создаем Excel с формулами
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = f"Аналитика_{filters['date_from']}_{filters['date_to']}"
            
            # 3. Сначала добавляем заголовок
            self._add_header(sheet, filters)
            
            # 4. Добавляем вертикальную часть с формулами (сначала с заглушками)
            # Она займет примерно 25 строк
            vertical_end_row = self._add_vertical_part_with_formulas(sheet, start_row=5)
            
            # 5. Добавляем отступ после вертикальной части
            horizontal_start_row = vertical_end_row + 3
            
            # 6. Добавляем горизонтальную часть (таблицу с товарами)
            # Определяем номера колонок для каждой метрики
            column_mapping = self._get_column_mapping_for_horizontal_table()
            
            # 7. Добавляем заголовки горизонтальной таблицы
            self._add_horizontal_table_headers(sheet, horizontal_start_row, column_mapping)
            
            # 8. Добавляем данные товаров
            data_start_row = horizontal_start_row + 1
            last_data_row = self._add_horizontal_table_data(
                sheet, raw_df, data_start_row, column_mapping
            )
            
            # 9. Добавляем строку "ИТОГО" в горизонтальной таблице
            total_row = last_data_row + 1
            self._add_horizontal_total_row(
                sheet, data_start_row, last_data_row, total_row, column_mapping
            )
            
            # 10. Теперь, когда знаем total_row, обновляем формулы в вертикальной части
            # Сначала определяем буквы колонок горизонтальной таблицы
            column_letters = self._map_column_indices_to_letters(column_mapping)
            
            # 11. Обновляем формулы в вертикальной части с правильными ссылками
            self._update_vertical_formulas(
                sheet, start_row=5, vertical_end_row=vertical_end_row,
                total_row=total_row, column_letters=column_letters
            )
            
            # 12. Настраиваем ширину столбцов
            self._adjust_column_widths(sheet)
            
            # 13. Сохраняем в BytesIO
            output = BytesIO()
            workbook.save(output)
            output.seek(0)
            
            logger.info(f"Сгенерирован отчет для tenant_id={tenant_id}, период: {filters['date_from']} - {filters['date_to']}")
            return output
            
        except Exception as e:
            logger.error(f"Ошибка генерации отчета: {e}", exc_info=True)
            raise
    
    def _add_header(self, sheet, filters: Dict[str, Any]):
        """Добавляет заголовок отчета"""
        # Основной заголовок
        sheet.cell(row=1, column=2, value="Отчет по аналитике Wildberries")
        sheet.cell(row=1, column=2).font = Font(bold=True, size=16)
        
        # Период отчета
        period_text = f"Период: {filters['date_from']} - {filters['date_to']}"
        sheet.cell(row=2, column=2, value=period_text)
        sheet.cell(row=2, column=2).font = Font(italic=True)
        
        # Дата генерации
        generation_date = datetime.now().strftime("%d.%m.%Y %H:%M")
        sheet.cell(row=3, column=2, value=f"Сгенерировано: {generation_date}")
        sheet.cell(row=3, column=2).font = Font(italic=True)
    
    def _get_report_data(self, filters: Dict, tenant_id: int) -> pd.DataFrame:
        """
        Запрос к БД для получения всех данных
        """
        try:
            # Выбираем только нужные колонки
            columns_to_select = list(self.COLUMN_MAPPING.keys())
            select_clause = ", ".join(columns_to_select)
            
            query = text(f"""
                SELECT 
                    {select_clause}
                FROM product_margins_month_v
                WHERE tenant_id = :tenant_id
                  AND period_month >= :date_from
                  AND period_month <= :date_to
                ORDER BY revenue DESC
            """)
            
            result = self.db.execute(query, {
                'tenant_id': tenant_id,
                'date_from': filters['date_from'],
                'date_to': filters['date_to']
            })
            
            # Создаем DataFrame с русскими названиями колонок
            df = pd.DataFrame(
                result.fetchall(), 
                columns=[self.COLUMN_MAPPING[col] for col in columns_to_select]
            )
            
            logger.info(f"Получено {len(df)} записей из БД")
            return df
            
        except Exception as e:
            logger.error(f"Ошибка получения данных из БД: {e}", exc_info=True)
            return pd.DataFrame()
    
    def _add_vertical_part_with_formulas(self, sheet, start_row: int) -> int:
        """
        Добавляет вертикальную часть отчета (ключевые показатели с формулами)
        Возвращает номер строки, где заканчивается вертикальная часть
        """
        # Заголовок вертикальной части
        title_cell = sheet.cell(row=start_row, column=2, value="Ключевые показатели")
        title_cell.font = Font(bold=True, size=14)
        
        # Заголовки таблицы
        sheet.cell(row=start_row + 1, column=2, value="Показатель")
        sheet.cell(row=start_row + 1, column=3, value="Значение")
        
        # Применяем стили к заголовкам
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col in [2, 3]:
            cell = sheet.cell(row=start_row + 1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = self._get_border()
        
        # Список показателей вертикальной части (22 показателя)
        # Формулы будут обновлены позже
        vertical_indicators = [
            ("Хранение", "={storage}"),
            ("Удержание", "={regular_deduction}"),
            ("Джем", "={dzhem_deduction}"),
            ("Реклама", "=0"),
            ("Общие расходы на рекламу", "=C8+C9+C10"),
            ("Доставка", "={delivery}"),
            ("Грузоперевозки", "=0"),
            ("Приемка", "={acceptance}"),
            ("Общие расходы на логистику", "=C12+C13+C14"),
            ("Маржа минус расходы", "{margin}-C7-C8-C9-C10-C13-C14"),
            ("Маржа на единицу товара", "=C16/{quantity}"),
            ("Магазинная наценка (от выручки)", "{margin}/{revenue}"),
            ("Магазинная наценка (от перечислений)", "{margin}/{payout}"),
            ("DRR", "=C8/{revenue}"),
            ("Зарплата фиксированная", "=60000"),
            ("Премия", "(C16+({regular_deduction}-{revenue}*0.1))*0.05"),
            ("Зарплата сотрудника", "=C21+C22"),
            ("Зарплата владельца", "=C16*0.1"),
            ("Маржа минус общая зарплата", "=C16-C23-C24"),
            ("Рентабельность", "=C25/{payout}"),
        ]
        
        # Добавляем показатели с шаблонными формулами
        current_row = start_row + 2
        for indicator_name, formula_template in vertical_indicators:
            # Название показателя
            cell_name = sheet.cell(row=current_row, column=2, value=indicator_name)
            cell_name.border = self._get_border()
            cell_name.alignment = Alignment(horizontal="left")
            
            # Формула-шаблон (будет обновлена позже)
            cell_value = sheet.cell(row=current_row, column=3, value=formula_template)
            cell_value.border = self._get_border()
            
            # Предварительное форматирование
            if any(word in indicator_name.lower() for word in ['%', 'процент', 'наценка', 'drr', 'рентабельность']):
                cell_value.number_format = '0.00%'
            else:
                cell_value.number_format = '#,##0.00'
            
            current_row += 1
        
        return current_row - 1  # Последняя заполненная строка
    
    def _get_column_mapping_for_horizontal_table(self) -> Dict[str, int]:
        """
        Создает маппинг метрик на номера колонок для горизонтальной таблицы
        """
        column_mapping = {}
        col_idx = 2
        
        # Проходим по всем колонкам в том порядке, в котором они должны отображаться
        for english_name, russian_name in self.COLUMN_MAPPING.items():
            # Добавляем русское название и его номер колонки
            column_mapping[russian_name] = col_idx
            
            # Также добавляем удобное имя для формул, если нужно
            if english_name == 'revenue':
                column_mapping['revenue'] = col_idx
            elif english_name == 'seller_payout':
                column_mapping['payout'] = col_idx
            elif english_name == 'margin':
                column_mapping['margin'] = col_idx
            elif english_name == 'quantity_sold':
                column_mapping['quantity'] = col_idx
            elif english_name == 'storage_fee':
                column_mapping['storage'] = col_idx
            elif english_name == 'regular_deduction':
                column_mapping['regular_deduction'] = col_idx
            elif english_name == 'dzhem_deduction':
                column_mapping['dzhem_deduction'] = col_idx
            elif english_name == 'delivery_rub':
                column_mapping['delivery'] = col_idx
            elif english_name == 'penalty':
                column_mapping['penalty'] = col_idx
            elif english_name == 'acceptance':
                column_mapping['acceptance'] = col_idx
            
            col_idx += 1
        
        return column_mapping
    
    def _add_horizontal_table_headers(self, sheet, start_row: int, column_mapping: Dict):
        """
        Добавляет заголовки горизонтальной таблицы
        """
        # Заголовок горизонтальной части
        sheet.cell(row=start_row, column=2, value="Детализация по товарам")
        sheet.cell(row=start_row, column=2).font = Font(bold=True, size=12)
        
        # Добавляем заголовки колонок в том же порядке, что и в COLUMN_MAPPING
        header_row = start_row + 1
        col_idx = 2
        
        for english_name, russian_name in self.COLUMN_MAPPING.items():
            cell = sheet.cell(row=header_row, column=col_idx, value=russian_name)
            self._apply_header_style(cell)
            col_idx += 1
    
    def _add_horizontal_table_data(self, sheet, df: pd.DataFrame, start_row: int, 
                                column_mapping: Dict) -> int:
        """
        Добавляет данные горизонтальной таблицы
        Возвращает номер последней строки с данными
        """
        # Создаем обратный маппинг: русское название -> номер колонки
        # Ищем русские названия, которые есть в DataFrame
        col_index_by_name = {}
        
        for key, value in column_mapping.items():
            # Берем только строковые ключи, которые являются русскими названиями
            if isinstance(key, str) and key in df.columns:
                col_index_by_name[key] = value
        
        # print(f"col_index_by_name: {col_index_by_name}")
        # print(f"DataFrame columns: {list(df.columns)}")
        
        # Добавляем данные
        current_row = start_row + 1
        
        # Проходим по строкам DataFrame
        for idx, row_data in df.iterrows():
            # print(f"Processing row {idx}")
            
            # Проходим по всем русским названиям колонок
            for russian_name in df.columns:
                if russian_name in col_index_by_name:
                    col_idx = col_index_by_name[russian_name]
                    value = row_data[russian_name]
                    
                    # Пропускаем NaN значения
                    if pd.isna(value):
                        continue
                    
                    cell = sheet.cell(row=current_row, column=col_idx, value=value)
                    self._apply_data_style(cell)
                    
                    # Форматирование числовых колонок
                    if '₽' in russian_name:
                        cell.number_format = '#,##0.00'
                    elif '%' in russian_name:
                        cell.number_format = '0.00%'
                    elif 'шт.' in russian_name.lower():
                        cell.number_format = '#,##0'
                else:
                    print(f"Warning: {russian_name} not found in col_index_by_name")
            
            current_row += 1
            # if idx < 2:  # Выводим первые несколько строк для отладки
            #     print(f"Row {idx} added at row {current_row-1}")
        
        return current_row - 1  # Последняя строка с данными
    
    def _add_horizontal_total_row(self, sheet, first_data_row: int, last_data_row: int,
                                  total_row: int, column_mapping: Dict):
        """
        Добавляет строку "ИТОГО" в горизонтальной таблице
        """
        # Заголовок "ИТОГО"
        cell = sheet.cell(row=total_row, column=2, value="ИТОГО")
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        cell.border = self._get_border()
        
        # Добавляем формулы СУММ для числовых колонок
        for col_idx in range(3, sheet.max_column + 1):
            col_letter = get_column_letter(col_idx)
            formula = f"=SUM({col_letter}{first_data_row + 1}:{col_letter}{last_data_row})"
            
            cell = sheet.cell(row=total_row, column=col_idx, value=formula)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            cell.border = self._get_border()
            
            # Форматирование
            header_cell = sheet.cell(row=first_data_row - 1, column=col_idx)
            if header_cell.value:
                header_text = str(header_cell.value)
                if '₽' in header_text:
                    cell.number_format = '#,##0.00'
                elif '%' in header_text:
                    cell.number_format = '0.00%'
                elif 'шт.' in header_text.lower():
                    cell.number_format = '#,##0'
    
    def _map_column_indices_to_letters(self, column_mapping: Dict) -> Dict[str, str]:
        """
        Преобразует номера колонок в буквы Excel для формул
        """
        letter_mapping = {}
        # print("col_map: ", column_mapping)
        # Преобразуем удобные имена в буквы колонок
        for key in ['revenue', 'payout', 'margin', 'quantity', 'storage', 
                    'regular_deduction', 'dzhem_deduction', 'delivery', 
                    'penalty', 'acceptance']:
            if key in column_mapping:
                col_idx = column_mapping[key]
                letter_mapping[key] = get_column_letter(col_idx)
        
        return letter_mapping
    
    def _update_vertical_formulas(self, sheet, start_row: int, vertical_end_row: int,
                                  total_row: int, column_letters: Dict[str, str]):
        """
        Обновляет формулы в вертикальной части с правильными ссылками
        """
        # print("col_let: ", column_letters)

        # Извлекаем буквы колонок
        rev = column_letters.get('revenue', 'E')
        pay = column_letters.get('payout', 'F')
        mar = column_letters.get('margin', 'O')
        qty = column_letters.get('quantity', 'D')
        sto = column_letters.get('storage', 'I')
        reg = column_letters.get('regular_deduction', 'J')
        dzhem = column_letters.get('dzhem_deduction', 'K')
        deliv = column_letters.get('delivery', 'L')
        penal = column_letters.get('penalty', 'M')
        accept = column_letters.get('acceptance', 'N')
        
        # Создаем словарь для замены плейсхолдеров
        replacements = {
            '{revenue}': f'{rev}{total_row}',
            '{payout}': f'{pay}{total_row}',
            '{margin}': f'{mar}{total_row}',
            '{quantity}': f'{qty}{total_row}',
            '{storage}': f'{sto}{total_row}',
            '{regular_deduction}': f'{reg}{total_row}',
            '{dzhem_deduction}': f'{dzhem}{total_row}',
            '{delivery}': f'{deliv}{total_row}',
            '{penalty}': f'{penal}{total_row}',
            '{acceptance}': f'{accept}{total_row}',
        }
        
        # Обновляем формулы в вертикальной части
        for row in range(start_row + 2, vertical_end_row + 1):
            cell = sheet.cell(row=row, column=3)  # Колонка C с формулами
            
            if cell.value and isinstance(cell.value, str):
                formula = str(cell.value)
                
                # Заменяем плейсхолдеры
                for placeholder, replacement in replacements.items():
                    formula = formula.replace(placeholder, replacement)
                
                # Заменяем {current_row} на номер текущей строки
                formula = formula.replace('{current_row}', str(row))
                
                # Если формула не начинается с '=', добавляем его
                if any(op in formula for op in ['+', '-', '*', '/', '(', ')']) and not formula.startswith('='):
                    formula = '=' + formula
                
                cell.value = formula
    
    def _adjust_column_widths(self, sheet):
        """Автоматически настраивает ширину столбцов"""
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min((max_length + 2) * 1.2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width
    
    def _get_border(self):
        """Создает границу для ячеек"""
        return Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
    
    def _apply_header_style(self, cell):
        """Применяет стиль к заголовкам"""
        cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = self._get_border()
    
    def _apply_data_style(self, cell):
        """Применяет стиль к данным"""
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = self._get_border()