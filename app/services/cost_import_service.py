# app/services/cost_import_service.py
"""
Массовый импорт себестоимостей из файла (xlsx/csv).

Формат: колонки sku, cost, start_date (name — необязательная, игнорируется).
Матчинг товаров — только по sku в рамках тенанта.

Семантика (согласована с владельцем продукта):
- запись с той же start_date, что в файле → стоимость ОБНОВЛЯЕТСЯ;
- новая запись с датой D → таймлайн себестоимостей товара пересегментируется:
  каждая запись получает end_date = start_date следующей − 1 день, последняя —
  NULL (это автоматически закрывает прежний открытый период, вставляет историю
  в середину и обрабатывает backfill раньше существующих записей);
- товар не найден по sku → строка-ошибка, записи не создаём;
- дубль внутри файла (sku + start_date) → побеждает последняя строка,
  предыдущая помечается ошибкой «переопределена».

После коммита обновляются материализованные view аналитики — себестоимость
участвует в расчёте маржи через supplier_reports_agg_mv (LATERAL-джойн).
"""
import csv
import io
import logging
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation
from typing import Dict, List, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy.orm import Session

from app.models.product import Product
from app.models.product_cost import ProductCost
from app.schemas.product_cost import (
    CostImportReport,
    CostImportRowReport,
    CostImportSummary,
)
from app.services.sync_service import refresh_analytics_materialized_views

logger = logging.getLogger(__name__)

MAX_ROWS = 5000
REQUIRED_COLUMNS = {"sku", "cost", "start_date"}
CREATED_BY = "file-import"
ALLOWED_EXTENSIONS = {".xlsx", ".csv"}

# Эпоха дат Excel (серийные номера дней)
_EXCEL_EPOCH = date(1899, 12, 30)


class CostImportError(Exception):
    """Ошибка формата файла — прокидывается в роутер как 400."""


class CostImportService:
    def __init__(self, db: Session):
        self.db = db

    # ------------------------------------------------------------------
    # ШАБЛОН (экспорт текущего состояния для раундтрипа)
    # ------------------------------------------------------------------

    def generate_template(self, tenant_id: int) -> io.BytesIO:
        """
        xlsx-шаблон с префиллом: все товары тенанта + текущая себестоимость.
        Пользователь правит/добавляет строки и загружает файл обратно.
        """
        products = (
            self.db.query(Product)
            .filter(Product.tenant_id == tenant_id)
            .order_by(Product.sku)
            .all()
        )
        product_ids = [p.id for p in products]

        # Текущие (открытые) себестоимости одним запросом
        current_costs: Dict[int, ProductCost] = {}
        if product_ids:
            open_costs = (
                self.db.query(ProductCost)
                .filter(
                    ProductCost.product_id.in_(product_ids),
                    ProductCost.end_date.is_(None),
                )
                .order_by(ProductCost.start_date.desc())
                .all()
            )
            for c in open_costs:
                current_costs.setdefault(c.product_id, c)

        wb = Workbook()
        ws = wb.active
        ws.title = "costs"

        headers = ["sku", "name", "cost", "start_date"]
        header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = header_fill

        for row_idx, product in enumerate(products, start=2):
            ws.cell(row=row_idx, column=1, value=product.sku)
            ws.cell(row=row_idx, column=2, value=product.name or "")
            current = current_costs.get(product.id)
            if current:
                ws.cell(row=row_idx, column=3, value=float(current.cost)).number_format = "0.00"
                ws.cell(row=row_idx, column=4, value=current.start_date).number_format = "DD.MM.YYYY"

        # Ширины колонок
        for col, width in zip("ABCD", (20, 45, 12, 14)):
            ws.column_dimensions[col].width = width

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    # ------------------------------------------------------------------
    # ПАРСИНГ
    # ------------------------------------------------------------------

    def _read_table(self, data: bytes, filename: str) -> List[List]:
        """Читает файл в таблицу (список строк-значений). Первая непустая строка — шапка."""
        if filename.endswith(".csv"):
            return self._read_csv(data)
        return self._read_xlsx(data)

    def _read_xlsx(self, data: bytes) -> List[List]:
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        try:
            ws = wb.active
            return [list(row) for row in ws.iter_rows(values_only=True)]
        finally:
            wb.close()

    def _read_csv(self, data: bytes) -> List[List]:
        # utf-8-sig снимает BOM (Excel); cp1251 — частый случай ручных выгрузок
        text = None
        for encoding in ("utf-8-sig", "cp1251"):
            try:
                text = data.decode(encoding)
                break
            except UnicodeDecodeError:
                continue
        if text is None:
            raise CostImportError("Не удалось определить кодировку CSV (ожидались UTF-8 или cp1251)")
        # sniff разделителя (Excel в ru-локали иногда кладёт ';')
        header_line = next((l for l in text.splitlines() if l.strip()), "")
        delimiter = ";" if header_line.count(";") > header_line.count(",") else ","
        return [row for row in csv.reader(io.StringIO(text), delimiter=delimiter)]

    def _parse_header(self, table: List[List]) -> Dict[str, int]:
        """Ищет шапку в первой непустой строке, возвращает имя колонки → индекс."""
        for row in table:
            values = [str(v).strip().lower() if v is not None else "" for v in row]
            if any(values):
                mapping = {name: idx for idx, name in enumerate(values) if name}
                missing = REQUIRED_COLUMNS - set(mapping)
                if missing:
                    raise CostImportError(
                        f"В файле нет обязательных колонок: {', '.join(sorted(missing))}. "
                        f"Найдены: {', '.join(sorted(mapping)) or '—'}"
                    )
                return mapping
        raise CostImportError("Файл пуст или не содержит строк")

    def _parse_cost(self, value) -> Optional[Decimal]:
        if value is None or (isinstance(value, str) and not value.strip()):
            return None
        if isinstance(value, Decimal):
            result = value
        elif isinstance(value, (int, float)):
            result = Decimal(str(value))
        else:
            cleaned = value.strip().replace("\xa0", "").replace(" ", "").replace(",", ".")
            try:
                result = Decimal(cleaned)
            except InvalidOperation:
                return None
        if result < 0 or result > Decimal("100000000"):
            return None
        return result

    def _parse_date(self, value) -> Optional[date]:
        if value is None or (isinstance(value, str) and not value.strip()):
            return None
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        if isinstance(value, (int, float)):
            # серийный номер даты Excel
            try:
                return _EXCEL_EPOCH + timedelta(days=int(value))
            except (OverflowError, ValueError):
                return None
        text = str(value).strip()
        for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y"):
            try:
                return datetime.strptime(text, fmt).date()
            except ValueError:
                continue
        return None

    def _parse_rows(self, table: List[List]) -> List[dict]:
        """Шапка + данные → список распарсенных строк с ошибками по строкам."""
        col_map = self._parse_header(table)
        header_row_idx = next(
            i for i, row in enumerate(table)
            if any(v is not None and str(v).strip() for v in row)
        )

        parsed: List[dict] = []
        data_rows = 0
        for row_num, row in enumerate(table[header_row_idx + 1:], start=header_row_idx + 2):
            if not any(v is not None and str(v).strip() for v in row):
                continue  # пустые строки пропускаем
            data_rows += 1
            if data_rows > MAX_ROWS:
                raise CostImportError(f"Файл содержит больше {MAX_ROWS} строк данных")

            def cell(name: str):
                idx = col_map.get(name)
                return row[idx] if idx is not None and idx < len(row) else None

            sku = str(cell("sku")).strip() if cell("sku") is not None else ""
            cost = self._parse_cost(cell("cost"))
            start_date = self._parse_date(cell("start_date"))

            error = None
            if not sku:
                error = "Не указан sku"
            elif cost is None:
                error = "Некорректная себестоимость (ожидается число ≥ 0)"
            elif start_date is None:
                error = "Некорректная дата (ожидается ДД.ММ.ГГГГ или ГГГГ-ММ-ДД)"

            parsed.append({
                "row_num": row_num,
                "sku": sku or None,
                "cost": cost,
                "start_date": start_date,
                "error": error,
            })
        return parsed

    # ------------------------------------------------------------------
    # ПРЕДПРОСМОТР / КОММИТ
    # ------------------------------------------------------------------

    def process(self, tenant_id: int, data: bytes, filename: str, dry_run: bool) -> CostImportReport:
        table = self._read_table(data, filename)
        parsed = self._parse_rows(table)

        # Товары тенанта одним запросом
        skus = [p["sku"] for p in parsed if p["sku"]]
        products_by_sku: Dict[str, Product] = {}
        if skus:
            for product in self.db.query(Product).filter(
                Product.tenant_id == tenant_id,
                Product.sku.in_(skus),
            ).all():
                products_by_sku.setdefault(product.sku, product)

        # Существующие себестоимости одним запросом
        product_ids = [p.id for p in products_by_sku.values()]
        costs_by_key: Dict[tuple, ProductCost] = {}
        costs_by_product: Dict[int, List[ProductCost]] = {}
        if product_ids:
            for c in self.db.query(ProductCost).filter(
                ProductCost.product_id.in_(product_ids)
            ).all():
                costs_by_key[(c.product_id, c.start_date)] = c
                costs_by_product.setdefault(c.product_id, []).append(c)

        # Дубли внутри файла: (sku, start_date) → побеждает ПОСЛЕДНЯЯ строка,
        # предыдущая помечается ошибкой «переопределена»
        by_key: Dict[tuple, dict] = {}
        for p in parsed:
            if p["error"] is None:
                key = (p["sku"], p["start_date"])
                prev = by_key.get(key)
                if prev is not None:
                    prev["error"] = f"Переопределена строкой {p['row_num']}"
                by_key[key] = p

        # Отчёт по строкам
        rows: List[CostImportRowReport] = []
        valid_rows: List[dict] = []
        for p in parsed:
            product = products_by_sku.get(p["sku"]) if p["sku"] else None
            if p["error"]:
                rows.append(CostImportRowReport(
                    row_num=p["row_num"], sku=p["sku"], cost=p["cost"],
                    start_date=p["start_date"], action="error", message=p["error"],
                ))
                continue
            if product is None:
                rows.append(CostImportRowReport(
                    row_num=p["row_num"], sku=p["sku"], cost=p["cost"],
                    start_date=p["start_date"], action="error",
                    message="Товар с таким sku не найден",
                ))
                continue

            existing = costs_by_key.get((product.id, p["start_date"]))
            action = "update" if existing else "create"
            rows.append(CostImportRowReport(
                row_num=p["row_num"], sku=p["sku"], product_name=product.name,
                cost=p["cost"], start_date=p["start_date"], action=action,
                message=(
                    "Обновит существующую запись с этой датой"
                    if existing else None
                ),
            ))
            valid_rows.append({**p, "product": product})

        # Сколько открытых периодов закроется (для сводки предпросмотра)
        closed_preview = 0
        seen_products = set()
        for vr in valid_rows:
            pid = vr["product"].id
            if costs_by_key.get((pid, vr["start_date"])) is None and pid not in seen_products:
                seen_products.add(pid)
                if any(c.end_date is None for c in costs_by_product.get(pid, [])):
                    closed_preview += 1

        summary = CostImportSummary(
            created=sum(1 for r in rows if r.action == "create"),
            updated=sum(1 for r in rows if r.action == "update"),
            closed_periods=closed_preview,
            errors=sum(1 for r in rows if r.action == "error"),
            total_rows=len(rows),
        )

        mv_refresh_ms = None
        if not dry_run and valid_rows:
            summary, mv_refresh_ms = self._apply(valid_rows, summary)

        return CostImportReport(dry_run=dry_run, rows=rows, summary=summary,
                                mv_refresh_ms=mv_refresh_ms)

    def _apply(self, valid_rows: List[dict], summary: CostImportSummary) -> tuple:
        """
        Применяет импорт в одной транзакции:
        update-строки обновляют cost существующих записей, create-строки
        добавляются, затем таймлайн каждого товара пересегментируется.
        """
        created = 0
        updated = 0
        closed_periods = 0

        try:
            by_product: Dict[int, List] = {}
            for vr in valid_rows:
                product = vr["product"]
                existing = (
                    self.db.query(ProductCost)
                    .filter(
                        ProductCost.product_id == product.id,
                        ProductCost.start_date == vr["start_date"],
                    )
                    .first()
                )
                if existing:
                    existing.cost = vr["cost"]
                    existing.created_by = CREATED_BY
                    updated += 1
                    record = existing
                else:
                    record = ProductCost(
                        product_id=product.id,
                        cost=vr["cost"],
                        start_date=vr["start_date"],
                        created_by=CREATED_BY,
                    )
                    self.db.add(record)
                    created += 1
                by_product.setdefault(product.id, []).append(record)

            # Пересегментация таймлайна: end_date = next.start_date − 1 день,
            # последней записи — NULL
            for product_id, records in by_product.items():
                all_records = (
                    self.db.query(ProductCost)
                    .filter(ProductCost.product_id == product_id)
                    .all()
                )
                # новые записи ещё не в сессии запросом — объединяем вручную
                known_ids = {r.id for r in all_records}
                merged = list(all_records) + [r for r in records if r.id not in known_ids]
                merged.sort(key=lambda r: r.start_date)
                for i, rec in enumerate(merged):
                    new_end = (
                        merged[i + 1].start_date - timedelta(days=1)
                        if i + 1 < len(merged) else None
                    )
                    if rec.end_date != new_end:
                        # Считаем закрытыми только ранее существовавшие открытые
                        # периоды (rec.id is not None); новые записи, сразу
                        # закрытые следующей датой импорта, — не в счёт
                        if (
                            rec.end_date is None
                            and new_end is not None
                            and rec.id is not None
                        ):
                            closed_periods += 1
                        rec.end_date = new_end

            self.db.commit()

            summary.created = created
            summary.updated = updated
            summary.closed_periods = closed_periods
        except Exception as e:
            self.db.rollback()
            logger.error(f"❌ Cost import failed: {e}", exc_info=True)
            raise

        # Мат.view пересчитываем ПОСЛЕ коммита; ошибка рефреша не валит импорт
        mv_refresh_ms = refresh_analytics_materialized_views()
        logger.info(
            f"🎯 Cost import applied: created={created}, updated={updated}, "
            f"closed={closed_periods}, mv_refresh={mv_refresh_ms}ms"
        )
        return summary, mv_refresh_ms
